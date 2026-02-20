"""
analyze.py  –  WCCP Portfolio Tracker
======================================
Reads WCCP_Master_Trade_Blotter.xlsx from ./data/
Produces:
  • /reports/portfolio_report_YYYY-MM-DD.pdf  – PDF with summary, tax lots, charts
  • /output/daily_pricing_YYYY-MM-DD.xlsx     – Pricing template with Bloomberg placeholders
"""

import os
import sys
import warnings
from collections import defaultdict
from datetime import date
from pathlib import Path

import matplotlib
matplotlib.use("Agg")                          # headless backend – no display needed
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.gridspec as gridspec

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ──────────────────────────────────────────────────────────
# Paths
# ──────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
DATA_DIR    = BASE_DIR / "data"
REPORTS_DIR = BASE_DIR / "reports"
OUTPUT_DIR  = BASE_DIR / "output"

TODAY       = date.today().isoformat()          # "YYYY-MM-DD"
BLOTTER_FILE = DATA_DIR / "WCCP_Master_Trade_Blotter.xlsx"
PDF_PATH    = REPORTS_DIR / f"portfolio_report_{TODAY}.pdf"
EXCEL_PATH  = OUTPUT_DIR  / f"daily_pricing_{TODAY}.xlsx"

# Create output dirs if missing
REPORTS_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ──────────────────────────────────────────────────────────
# Colour palette
# ──────────────────────────────────────────────────────────
NAVY   = "#1B3A5C"
GOLD   = "#B8960C"
LIGHT  = "#F5F7FA"
WHITE  = "#FFFFFF"
GREY   = "#6B7280"
RED    = "#DC2626"
GREEN  = "#16A34A"
PIE_COLORS = [
    "#1B3A5C", "#2563EB", "#3B82F6", "#60A5FA",
    "#93C5FD", "#BFDBFE", "#DBEAFE", "#EFF6FF",
    "#B8960C", "#D97706", "#F59E0B", "#FCD34D",
]

# ══════════════════════════════════════════════════════════
# 1. DATA LOADING
# ══════════════════════════════════════════════════════════

def load_blotter(path: Path) -> pd.DataFrame:
    """Load and normalise the trade blotter."""
    if not path.exists():
        sys.exit(
            f"\n[ERROR] Trade blotter not found: {path}\n"
            f"        Place WCCP_Master_Trade_Blotter.xlsx in the ./data/ folder.\n"
        )

    df = pd.read_excel(path, sheet_name="Trade Blotter")

    # Normalise column names
    df.columns = [c.strip() for c in df.columns]

    required = {
        "Date", "Security", "Ticker", "Type",
        "Transaction Type", "Quantity", "Price",
        "Fund Consideration", "Currency", "Entity", "AccountShortName",
    }
    missing = required - set(df.columns)
    if missing:
        sys.exit(f"\n[ERROR] Missing columns in blotter: {missing}\n")

    # Type coercion
    df["Date"]     = pd.to_datetime(df["Date"])
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
    df["Price"]    = pd.to_numeric(df["Price"],    errors="coerce").fillna(0)
    df["Fund Consideration"] = pd.to_numeric(
        df["Fund Consideration"], errors="coerce"
    ).fillna(0)
    df["Transaction Type"] = df["Transaction Type"].str.strip().str.title()  # Buy / Sell
    df = df.sort_values("Date").reset_index(drop=True)

    return df


# ══════════════════════════════════════════════════════════
# 2. FIFO TAX-LOT MATCHING
# ══════════════════════════════════════════════════════════

def fifo_match(df: pd.DataFrame):
    """
    FIFO lot matching per Ticker + Entity.

    Returns
    -------
    open_lots   : list of dicts  (one per surviving buy lot)
    realized_pl : list of dicts  (one per matched sell)
    """
    # Group trades by (Ticker, Entity)
    groups = defaultdict(list)
    for _, row in df.iterrows():
        key = (row["Ticker"], row["Entity"])
        groups[key].append(row)

    open_lots    = []   # remaining open tax lots
    realized_pl  = []   # matched sell records

    for (ticker, entity), trades in groups.items():
        buy_queue = []   # list of [date, security, qty_remaining, price, entity]

        for trade in trades:
            qty   = abs(float(trade["Quantity"]))
            price = float(trade["Price"])
            txn   = trade["Transaction Type"]

            if txn == "Buy":
                buy_queue.append({
                    "date":     trade["Date"],
                    "security": trade["Security"],
                    "ticker":   ticker,
                    "entity":   entity,
                    "account":  trade["AccountShortName"],
                    "currency": trade["Currency"],
                    "qty_rem":  qty,
                    "cost_px":  price,
                })

            elif txn == "Sell":
                qty_to_sell = qty
                while qty_to_sell > 0 and buy_queue:
                    lot = buy_queue[0]
                    matched = min(lot["qty_rem"], qty_to_sell)

                    realized_pl.append({
                        "ticker":      ticker,
                        "entity":      entity,
                        "security":    lot["security"],
                        "buy_date":    lot["date"],
                        "sell_date":   trade["Date"],
                        "qty":         matched,
                        "cost_px":     lot["cost_px"],
                        "sell_px":     price,
                        "realized_pl": matched * (price - lot["cost_px"]),
                    })

                    lot["qty_rem"] -= matched
                    qty_to_sell    -= matched

                    if lot["qty_rem"] <= 1e-9:
                        buy_queue.pop(0)

                # If sells exceed buys (short / data issue), ignore remainder
                if qty_to_sell > 1e-4:
                    print(
                        f"  [WARN] Excess sell qty {qty_to_sell:.4f} for "
                        f"{ticker} / {entity} on {trade['Date'].date()} — skipped."
                    )

        # Remaining open lots
        for lot in buy_queue:
            if lot["qty_rem"] > 1e-9:
                open_lots.append(lot)

    return open_lots, realized_pl


# ══════════════════════════════════════════════════════════
# 3. POSITION SUMMARY
# ══════════════════════════════════════════════════════════

def build_positions(open_lots: list) -> pd.DataFrame:
    """
    Aggregate open lots into a position-level summary per (Ticker, Entity).
    """
    rows = []
    # group by ticker + entity
    grouped = defaultdict(list)
    for lot in open_lots:
        grouped[(lot["ticker"], lot["entity"])].append(lot)

    for (ticker, entity), lots in grouped.items():
        total_qty  = sum(l["qty_rem"] for l in lots)
        total_cost = sum(l["qty_rem"] * l["cost_px"] for l in lots)
        avg_cost   = total_cost / total_qty if total_qty else 0

        rows.append({
            "Security":        lots[0]["security"],
            "Ticker":          ticker,
            "Entity":          entity,
            "Account":         lots[0]["account"],
            "Currency":        lots[0]["currency"],
            "Total Shares":    total_qty,
            "Avg Cost":        avg_cost,
            "Total Cost Basis": total_cost,
        })

    df = pd.DataFrame(rows).sort_values(
        ["Entity", "Total Cost Basis"], ascending=[True, False]
    ).reset_index(drop=True)
    return df


# ══════════════════════════════════════════════════════════
# 4. PDF REPORT
# ══════════════════════════════════════════════════════════

def _fmt_num(val, decimals=2, prefix="$"):
    """Format a number with commas and optional prefix."""
    if pd.isna(val):
        return "—"
    return f"{prefix}{val:,.{decimals}f}" if prefix else f"{val:,.{decimals}f}"


def _draw_table(ax, df_table, col_widths=None, row_height=0.06,
                header_color=NAVY, header_font_color=WHITE,
                alt_color="#EBF0F5", font_size=7):
    """
    Draw a pandas DataFrame as a formatted table inside a matplotlib Axes.
    ax should already have axis('off').
    Returns the final y position (bottom of table).
    """
    ax.axis("off")
    cols   = list(df_table.columns)
    n_cols = len(cols)
    n_rows = len(df_table)

    if col_widths is None:
        col_widths = [1 / n_cols] * n_cols

    # Normalise widths to sum=1
    total_w = sum(col_widths)
    col_widths = [w / total_w for w in col_widths]

    x_starts = []
    x = 0
    for w in col_widths:
        x_starts.append(x)
        x += w

    y = 1.0

    # Header row
    for i, col in enumerate(cols):
        ax.add_patch(plt.Rectangle(
            (x_starts[i], y - row_height), col_widths[i], row_height,
            transform=ax.transAxes, color=header_color, clip_on=False
        ))
        ax.text(
            x_starts[i] + col_widths[i] / 2,
            y - row_height / 2,
            col, transform=ax.transAxes,
            ha="center", va="center",
            fontsize=font_size, fontweight="bold",
            color=header_font_color
        )
    y -= row_height

    # Data rows
    for r_idx, (_, row) in enumerate(df_table.iterrows()):
        fill = alt_color if r_idx % 2 == 0 else WHITE
        for i, col in enumerate(cols):
            ax.add_patch(plt.Rectangle(
                (x_starts[i], y - row_height), col_widths[i], row_height,
                transform=ax.transAxes, color=fill, clip_on=False
            ))
            # Right-align numbers, left-align text
            val = str(row[col])
            ha  = "right" if val.replace(",", "").replace("$", "").replace(
                "-", "").replace(".", "").replace("(", "").replace(")", "").isdigit() else "left"
            ax.text(
                x_starts[i] + (col_widths[i] * 0.95 if ha == "right" else col_widths[i] * 0.05),
                y - row_height / 2,
                val, transform=ax.transAxes,
                ha=ha, va="center",
                fontsize=font_size, color="#111827"
            )
        y -= row_height

    return y


def _page_header(fig, title_text, subtitle=""):
    """Add a navy header band at the top of a figure."""
    # Full-width header bar
    header_ax = fig.add_axes([0, 0.93, 1, 0.07])
    header_ax.set_facecolor(NAVY)
    header_ax.axis("off")
    header_ax.text(
        0.03, 0.55, "WCCP Portfolio Report",
        transform=header_ax.transAxes,
        fontsize=11, fontweight="bold", color=WHITE, va="center"
    )
    header_ax.text(
        0.03, 0.15, f"Generated: {TODAY}",
        transform=header_ax.transAxes,
        fontsize=7, color="#93C5FD", va="center"
    )
    header_ax.text(
        0.97, 0.55, title_text,
        transform=header_ax.transAxes,
        fontsize=10, fontweight="bold", color=GOLD,
        va="center", ha="right"
    )
    if subtitle:
        header_ax.text(
            0.97, 0.15, subtitle,
            transform=header_ax.transAxes,
            fontsize=7, color="#93C5FD", va="center", ha="right"
        )


def generate_pdf(
    positions: pd.DataFrame,
    open_lots: list,
    realized_pl: list,
    blotter: pd.DataFrame,
):
    print(f"  Generating PDF → {PDF_PATH}")

    total_cost    = positions["Total Cost Basis"].sum()
    total_pos     = len(positions)
    realized_total = sum(r["realized_pl"] for r in realized_pl)

    # Unique entities
    entities = sorted(positions["Entity"].dropna().unique())

    with PdfPages(str(PDF_PATH)) as pdf:

        # ── PAGE 1: Cover / Summary ──────────────────────────────────
        fig = plt.figure(figsize=(8.5, 11))
        fig.patch.set_facecolor(LIGHT)
        _page_header(fig, "Portfolio Summary")

        # Hero stats box
        stats_ax = fig.add_axes([0.05, 0.73, 0.9, 0.18])
        stats_ax.set_facecolor(WHITE)
        stats_ax.axis("off")
        for spine in stats_ax.spines.values():
            spine.set_visible(False)

        stats = [
            ("Total Positions",    str(total_pos)),
            ("Total Cost Basis",   f"${total_cost:,.2f}"),
            ("Realized P&L",       f"${realized_total:,.2f}"),
            ("Unrealized P&L",     "Pending Prices"),
        ]
        for i, (label, value) in enumerate(stats):
            x = 0.05 + i * 0.24
            color = GREEN if "P&L" in label and realized_total >= 0 else (
                RED if "P&L" in label and realized_total < 0 else NAVY
            )
            if "Unrealized" in label:
                color = GREY
            stats_ax.text(x, 0.7,  value, fontsize=14, fontweight="bold",
                           color=color, transform=stats_ax.transAxes)
            stats_ax.text(x, 0.3,  label, fontsize=8, color=GREY,
                           transform=stats_ax.transAxes)

        # Position summary table
        tbl_ax = fig.add_axes([0.05, 0.28, 0.9, 0.43])
        pos_display = positions[[
            "Security", "Ticker", "Entity",
            "Total Shares", "Avg Cost", "Total Cost Basis"
        ]].copy()
        pos_display["Total Shares"]    = pos_display["Total Shares"].map(lambda v: f"{v:,.4f}")
        pos_display["Avg Cost"]        = pos_display["Avg Cost"].map(lambda v: f"${v:,.4f}")
        pos_display["Total Cost Basis"] = pos_display["Total Cost Basis"].map(lambda v: f"${v:,.2f}")
        pos_display.columns = [
            "Security", "Ticker", "Entity",
            "Shares", "Avg Cost", "Cost Basis"
        ]

        _draw_table(
            tbl_ax, pos_display,
            col_widths=[2.5, 1.8, 1.5, 1.2, 1.2, 1.5],
            font_size=7
        )

        tbl_ax.text(
            0, -0.04, "Position Summary – Open Lots (FIFO)",
            transform=tbl_ax.transAxes,
            fontsize=9, fontweight="bold", color=NAVY
        )

        # Footer
        fig.text(0.5, 0.02, "CONFIDENTIAL  |  WCCP Internal Use Only",
                 ha="center", fontsize=7, color=GREY)
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        # ── PAGE 2: Allocation Charts ────────────────────────────────
        fig = plt.figure(figsize=(8.5, 11))
        fig.patch.set_facecolor(LIGHT)
        _page_header(fig, "Allocation Analysis")

        # Pie by position (cost basis weight)
        pie_ax = fig.add_axes([0.04, 0.52, 0.44, 0.38])
        pie_data    = positions.groupby("Ticker")["Total Cost Basis"].sum()
        pie_labels  = pie_data.index.tolist()
        pie_values  = pie_data.values
        colors_used = [PIE_COLORS[i % len(PIE_COLORS)] for i in range(len(pie_labels))]
        wedges, texts, autotexts = pie_ax.pie(
            pie_values, labels=None,
            autopct=lambda p: f"{p:.1f}%" if p > 3 else "",
            colors=colors_used, startangle=140,
            pctdistance=0.8,
            wedgeprops={"edgecolor": WHITE, "linewidth": 1.2},
        )
        for at in autotexts:
            at.set_fontsize(6)
        pie_ax.set_title("By Position (Cost Basis Weight)", fontsize=9,
                          fontweight="bold", color=NAVY, pad=6)

        # Legend
        legend_patches = [
            mpatches.Patch(color=colors_used[i], label=f"{pie_labels[i]} ({pie_values[i]/total_cost*100:.1f}%)")
            for i in range(len(pie_labels))
        ]
        pie_ax.legend(handles=legend_patches, loc="lower left",
                       bbox_to_anchor=(-0.2, -0.35),
                       fontsize=6, frameon=False)

        # Pie by entity
        ent_ax = fig.add_axes([0.54, 0.52, 0.44, 0.38])
        ent_data   = positions.groupby("Entity")["Total Cost Basis"].sum()
        ent_labels = ent_data.index.tolist()
        ent_values = ent_data.values
        ent_colors = [PIE_COLORS[(4 + i) % len(PIE_COLORS)] for i in range(len(ent_labels))]
        ent_ax.pie(
            ent_values, labels=None,
            autopct=lambda p: f"{p:.1f}%" if p > 3 else "",
            colors=ent_colors, startangle=140,
            pctdistance=0.8,
            wedgeprops={"edgecolor": WHITE, "linewidth": 1.2},
        )
        ent_ax.set_title("By Entity (Cost Basis Weight)", fontsize=9,
                          fontweight="bold", color=NAVY, pad=6)
        ent_legend = [
            mpatches.Patch(color=ent_colors[i], label=f"{ent_labels[i]} ({ent_values[i]/total_cost*100:.1f}%)")
            for i in range(len(ent_labels))
        ]
        ent_ax.legend(handles=ent_legend, loc="lower left",
                       bbox_to_anchor=(-0.2, -0.35),
                       fontsize=6, frameon=False)

        # Bar chart: cost basis by ticker
        bar_ax = fig.add_axes([0.08, 0.10, 0.88, 0.36])
        tickers_sorted = pie_data.sort_values(ascending=False)
        bars = bar_ax.barh(
            tickers_sorted.index[::-1],
            tickers_sorted.values[::-1],
            color=NAVY, edgecolor=WHITE, linewidth=0.5
        )
        bar_ax.set_xlabel("Cost Basis ($)", fontsize=8, color=GREY)
        bar_ax.set_title("Cost Basis by Ticker", fontsize=9,
                          fontweight="bold", color=NAVY)
        bar_ax.tick_params(labelsize=7)
        bar_ax.xaxis.set_major_formatter(
            matplotlib.ticker.FuncFormatter(lambda v, _: f"${v:,.0f}")
        )
        bar_ax.spines["top"].set_visible(False)
        bar_ax.spines["right"].set_visible(False)
        for bar, val in zip(bars, tickers_sorted.values[::-1]):
            bar_ax.text(
                val * 1.01, bar.get_y() + bar.get_height() / 2,
                f"${val:,.0f}", va="center", fontsize=6, color=NAVY
            )

        fig.text(0.5, 0.02, "CONFIDENTIAL  |  WCCP Internal Use Only",
                 ha="center", fontsize=7, color=GREY)
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        # ── PAGE 3: Realized vs Unrealized P&L ──────────────────────
        fig = plt.figure(figsize=(8.5, 11))
        fig.patch.set_facecolor(LIGHT)
        _page_header(fig, "P&L Analysis")

        # Realized P&L table
        if realized_pl:
            rpl_df = pd.DataFrame(realized_pl)
            rpl_display = rpl_df[[
                "ticker", "entity", "buy_date", "sell_date",
                "qty", "cost_px", "sell_px", "realized_pl"
            ]].copy()
            rpl_display["buy_date"]    = rpl_display["buy_date"].dt.strftime("%Y-%m-%d")
            rpl_display["sell_date"]   = rpl_display["sell_date"].dt.strftime("%Y-%m-%d")
            rpl_display["qty"]         = rpl_display["qty"].map(lambda v: f"{v:,.4f}")
            rpl_display["cost_px"]     = rpl_display["cost_px"].map(lambda v: f"${v:,.4f}")
            rpl_display["sell_px"]     = rpl_display["sell_px"].map(lambda v: f"${v:,.4f}")
            rpl_display["realized_pl"] = rpl_display["realized_pl"].map(lambda v: f"${v:,.2f}")
            rpl_display.columns = [
                "Ticker", "Entity", "Buy Date", "Sell Date",
                "Qty", "Cost Px", "Sell Px", "Realized P&L"
            ]

            rpl_ax = fig.add_axes([0.04, 0.48, 0.92, 0.42])
            _draw_table(
                rpl_ax, rpl_display,
                col_widths=[1.5, 1.4, 1.2, 1.2, 1.0, 1.1, 1.1, 1.3],
                font_size=7
            )
            rpl_ax.text(
                0, -0.04, f"Realized P&L Summary  |  Total: ${realized_total:,.2f}",
                transform=rpl_ax.transAxes,
                fontsize=9, fontweight="bold",
                color=GREEN if realized_total >= 0 else RED
            )
        else:
            no_ax = fig.add_axes([0.1, 0.55, 0.8, 0.3])
            no_ax.axis("off")
            no_ax.text(0.5, 0.5, "No realized transactions found.",
                        ha="center", va="center", fontsize=10, color=GREY,
                        transform=no_ax.transAxes)

        # Unrealized P&L placeholder notice
        unr_ax = fig.add_axes([0.04, 0.22, 0.92, 0.22])
        unr_ax.set_facecolor("#FEF3C7")
        unr_ax.axis("off")
        for spine in unr_ax.spines.values():
            spine.set_visible(False)
        unr_ax.text(
            0.03, 0.82, "Unrealized P&L",
            transform=unr_ax.transAxes,
            fontsize=10, fontweight="bold", color="#92400E"
        )
        unr_ax.text(
            0.03, 0.55,
            "Current market prices are required to calculate unrealized P&L.",
            transform=unr_ax.transAxes, fontsize=8, color="#78350F"
        )
        unr_ax.text(
            0.03, 0.35,
            "Use the accompanying daily_pricing_YYYY-MM-DD.xlsx file:",
            transform=unr_ax.transAxes, fontsize=8, color="#78350F"
        )
        unr_ax.text(
            0.03, 0.15,
            '1. Open the Excel file  2. Enter current prices or Bloomberg formula =BDP("TICKER","PX_LAST")'
            '  3. Market Value & P&L columns auto-calculate.',
            transform=unr_ax.transAxes, fontsize=7.5, color="#78350F"
        )

        fig.text(0.5, 0.02, "CONFIDENTIAL  |  WCCP Internal Use Only",
                 ha="center", fontsize=7, color=GREY)
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)

        # ── PAGES 4+: Tax Lot Detail ─────────────────────────────────
        lots_df = pd.DataFrame(open_lots)
        if lots_df.empty:
            pass
        else:
            tickers_in_lots = lots_df["ticker"].unique()
            for ticker in sorted(tickers_in_lots):
                ticker_lots = lots_df[lots_df["ticker"] == ticker].copy()

                fig = plt.figure(figsize=(8.5, 11))
                fig.patch.set_facecolor(LIGHT)
                sec_name = ticker_lots.iloc[0]["security"]
                _page_header(fig, f"Tax Lot Detail", subtitle=f"{sec_name}  |  {ticker}")

                # Lot table
                lot_display = ticker_lots[[
                    "date", "entity", "account",
                    "qty_rem", "cost_px"
                ]].copy()
                lot_display["date"]     = lot_display["date"].dt.strftime("%Y-%m-%d")
                lot_display["cost_basis"] = lot_display["qty_rem"] * lot_display["cost_px"]
                lot_display["qty_rem"]  = lot_display["qty_rem"].map(lambda v: f"{v:,.4f}")
                lot_display["cost_px"]  = lot_display["cost_px"].map(lambda v: f"${v:,.4f}")
                lot_display["cost_basis"] = lot_display["cost_basis"].map(lambda v: f"${v:,.2f}")

                lot_display.columns = [
                    "Buy Date", "Entity", "Account",
                    "Open Qty", "Cost Price", "Cost Basis"
                ]

                lot_ax = fig.add_axes([0.04, 0.55, 0.92, 0.35])
                _draw_table(
                    lot_ax, lot_display,
                    col_widths=[1.5, 1.5, 1.5, 1.3, 1.3, 1.3],
                    font_size=8
                )

                # Summary stats for this ticker
                total_qty_t  = ticker_lots["qty_rem"].sum()
                total_cost_t = (ticker_lots["qty_rem"] * ticker_lots["cost_px"]).sum()
                avg_cost_t   = total_cost_t / total_qty_t if total_qty_t else 0

                summary_ax = fig.add_axes([0.04, 0.40, 0.92, 0.13])
                summary_ax.set_facecolor(WHITE)
                summary_ax.axis("off")
                summary_stats = [
                    ("Total Open Qty",  f"{total_qty_t:,.4f}"),
                    ("Avg Cost Price",  f"${avg_cost_t:,.4f}"),
                    ("Total Cost Basis", f"${total_cost_t:,.2f}"),
                    ("Current Price",   "— (enter in Excel)"),
                    ("Unrealized P&L",  "— (enter price)"),
                ]
                for i, (lbl, val) in enumerate(summary_stats):
                    x = 0.01 + i * 0.20
                    color = GREY if "enter" in val else NAVY
                    summary_ax.text(x, 0.7, val, fontsize=10, fontweight="bold",
                                     color=color, transform=summary_ax.transAxes)
                    summary_ax.text(x, 0.2, lbl, fontsize=7, color=GREY,
                                     transform=summary_ax.transAxes)

                fig.text(0.5, 0.02, "CONFIDENTIAL  |  WCCP Internal Use Only",
                         ha="center", fontsize=7, color=GREY)
                pdf.savefig(fig, bbox_inches="tight")
                plt.close(fig)

        # ── Entity Breakdown Pages ───────────────────────────────────
        for entity in entities:
            ent_pos = positions[positions["Entity"] == entity].copy()
            if ent_pos.empty:
                continue

            fig = plt.figure(figsize=(8.5, 11))
            fig.patch.set_facecolor(LIGHT)
            _page_header(fig, "Entity Breakdown", subtitle=entity)

            ent_cost = ent_pos["Total Cost Basis"].sum()

            # Mini pie
            mini_pie_ax = fig.add_axes([0.55, 0.62, 0.40, 0.28])
            if len(ent_pos) > 1:
                ep_vals   = ent_pos["Total Cost Basis"].values
                ep_labels = ent_pos["Ticker"].values
                mini_pie_ax.pie(
                    ep_vals, labels=None,
                    autopct=lambda p: f"{p:.1f}%" if p > 5 else "",
                    colors=[PIE_COLORS[i % len(PIE_COLORS)] for i in range(len(ep_vals))],
                    wedgeprops={"edgecolor": WHITE, "linewidth": 1},
                    startangle=140, pctdistance=0.8
                )
                mini_pie_ax.set_title("Weight by Cost Basis", fontsize=8,
                                       fontweight="bold", color=NAVY)
            else:
                mini_pie_ax.axis("off")
                mini_pie_ax.text(0.5, 0.5, "Single position",
                                  ha="center", va="center", fontsize=8, color=GREY)

            # Entity position table
            ent_tbl_ax = fig.add_axes([0.04, 0.36, 0.90, 0.52])
            ent_display = ent_pos[[
                "Security", "Ticker", "Total Shares", "Avg Cost", "Total Cost Basis"
            ]].copy()
            ent_display["Weight"] = (
                ent_display["Total Cost Basis"] / ent_cost * 100
            ).map(lambda v: f"{v:.1f}%")
            ent_display["Total Shares"]    = ent_display["Total Shares"].map(lambda v: f"{v:,.4f}")
            ent_display["Avg Cost"]        = ent_display["Avg Cost"].map(lambda v: f"${v:,.4f}")
            ent_display["Total Cost Basis"] = ent_display["Total Cost Basis"].map(lambda v: f"${v:,.2f}")
            ent_display.columns = [
                "Security", "Ticker", "Shares", "Avg Cost", "Cost Basis", "Weight"
            ]
            _draw_table(
                ent_tbl_ax, ent_display,
                col_widths=[2.5, 1.8, 1.3, 1.3, 1.5, 0.9],
                font_size=8
            )

            # Summary line
            summary_line_ax = fig.add_axes([0.04, 0.28, 0.92, 0.07])
            summary_line_ax.set_facecolor(NAVY)
            summary_line_ax.axis("off")
            summary_line_ax.text(
                0.02, 0.5,
                f"{entity}  |  Positions: {len(ent_pos)}  |  "
                f"Total Cost Basis: ${ent_cost:,.2f}  |  "
                f"Portfolio Weight: {ent_cost/total_cost*100:.1f}%",
                transform=summary_line_ax.transAxes,
                fontsize=8, fontweight="bold", color=WHITE, va="center"
            )

            fig.text(0.5, 0.02, "CONFIDENTIAL  |  WCCP Internal Use Only",
                     ha="center", fontsize=7, color=GREY)
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)

    print(f"  PDF saved → {PDF_PATH}")


# ══════════════════════════════════════════════════════════
# 5. DAILY PRICING EXCEL
# ══════════════════════════════════════════════════════════

def generate_excel(positions: pd.DataFrame):
    print(f"  Generating Excel → {EXCEL_PATH}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Pricing"

    # ── Styles ──────────────────────────────────────────────
    navy_fill   = PatternFill("solid", fgColor="1B3A5C")
    gold_fill   = PatternFill("solid", fgColor="B8960C")
    yellow_fill = PatternFill("solid", fgColor="FEF9C3")
    light_fill  = PatternFill("solid", fgColor="EBF0F5")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")
    green_fill  = PatternFill("solid", fgColor="DCFCE7")
    red_fill    = PatternFill("solid", fgColor="FEE2E2")

    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    subhdr_font  = Font(name="Calibri", bold=True, color="1B3A5C", size=9)
    data_font    = Font(name="Calibri", size=9)
    formula_font = Font(name="Calibri", size=9, color="1D4ED8")  # blue for formulas

    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align  = Alignment(horizontal="right",  vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")

    # ── Title row ────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    ws["A1"] = f"WCCP Daily Portfolio Pricing Sheet  –  {TODAY}"
    ws["A1"].font   = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill   = navy_fill
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    ws["A2"] = (
        'Instructions: Enter current prices in Column G, or use Bloomberg formula  '
        '=BDP("TICKER","PX_LAST")  — Market Value, P&L, and % Return auto-calculate.'
    )
    ws["A2"].font      = Font(name="Calibri", italic=True, size=8, color="78350F")
    ws["A2"].fill      = PatternFill("solid", fgColor="FEF3C7")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    # ── Column headers (row 3) ───────────────────────────────
    headers = [
        "Security", "Ticker", "Entity", "Account",
        "Total Shares", "Avg Cost", "Current Price",
        "Market Value", "Total Cost Basis", "Unrealized P&L",
        "% Return", "Currency"
    ]
    col_widths = [28, 20, 18, 18, 13, 13, 14, 16, 16, 16, 10, 10]

    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font      = header_font
        cell.fill      = navy_fill
        cell.alignment = center_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 20

    # ── Data rows ────────────────────────────────────────────
    # Column letter map
    # A=Security B=Ticker C=Entity D=Account
    # E=Total Shares F=Avg Cost G=Current Price
    # H=Market Value I=Total Cost Basis J=Unrealized P&L K=% Return L=Currency

    data_start_row = 4
    for r_offset, (_, pos) in enumerate(positions.iterrows()):
        row = data_start_row + r_offset
        alt_fill = light_fill if r_offset % 2 == 0 else white_fill

        def cell_set(col, value, num_fmt=None, font=None, fill=None, align=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = font  or data_font
            c.fill      = fill  or alt_fill
            c.alignment = align or left_align
            c.border    = border
            if num_fmt:
                c.number_format = num_fmt
            return c

        ticker_raw  = pos["Ticker"]                    # e.g. "LEU US Equity"
        ticker_bbg  = ticker_raw                        # keep Bloomberg format

        cell_set(1,  pos["Security"])
        cell_set(2,  ticker_raw)
        cell_set(3,  pos["Entity"])
        cell_set(4,  pos.get("Account", ""))
        cell_set(5,  pos["Total Shares"],    num_fmt="#,##0.0000", align=right_align)
        cell_set(6,  pos["Avg Cost"],        num_fmt='$#,##0.0000', align=right_align)

        # Current Price – input cell with comment
        price_cell = ws.cell(row=row, column=7, value=None)
        price_cell.font      = formula_font
        price_cell.fill      = yellow_fill
        price_cell.alignment = right_align
        price_cell.border    = border
        price_cell.number_format = '$#,##0.0000'
        comment_text = (
            f'Enter current price manually, or use Bloomberg formula:\n'
            f'=BDP("{ticker_bbg}","PX_LAST")'
        )
        price_cell.comment = Comment(comment_text, "WCCP System")

        e_col  = get_column_letter(5)   # Total Shares
        f_col  = get_column_letter(6)   # Avg Cost
        g_col  = get_column_letter(7)   # Current Price
        h_col  = get_column_letter(8)   # Market Value
        i_col  = get_column_letter(9)   # Total Cost Basis

        # Market Value = Shares * Current Price (formula, shown in blue)
        mv_cell = ws.cell(row=row, column=8,
                           value=f"=IF({g_col}{row}=\"\",\"\",{e_col}{row}*{g_col}{row})")
        mv_cell.font         = formula_font
        mv_cell.fill         = alt_fill
        mv_cell.alignment    = right_align
        mv_cell.border       = border
        mv_cell.number_format = '$#,##0.00'

        # Total Cost Basis (static value)
        cell_set(9, pos["Total Cost Basis"], num_fmt='$#,##0.00', align=right_align)

        # Unrealized P&L = Market Value - Total Cost Basis
        upl_cell = ws.cell(row=row, column=10,
                            value=f"=IF({g_col}{row}=\"\",\"\",{h_col}{row}-{i_col}{row})")
        upl_cell.font         = formula_font
        upl_cell.fill         = alt_fill
        upl_cell.alignment    = right_align
        upl_cell.border       = border
        upl_cell.number_format = '$#,##0.00'

        # % Return = Unrealized P&L / Total Cost Basis
        j_col = get_column_letter(10)
        ret_cell = ws.cell(row=row, column=11,
                           value=f"=IF({g_col}{row}=\"\",\"\",{j_col}{row}/{i_col}{row})")
        ret_cell.font         = formula_font
        ret_cell.fill         = alt_fill
        ret_cell.alignment    = right_align
        ret_cell.border       = border
        ret_cell.number_format = '0.00%'

        cell_set(12, pos.get("Currency", "USD"), align=center_align)

        ws.row_dimensions[row].height = 16

    # ── Totals row ───────────────────────────────────────────
    last_data_row = data_start_row + len(positions) - 1
    totals_row    = last_data_row + 1

    ws.row_dimensions[totals_row].height = 18

    total_lbl = ws.cell(row=totals_row, column=1, value="TOTAL")
    total_lbl.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    total_lbl.fill      = navy_fill
    total_lbl.alignment = center_align
    total_lbl.border    = border

    # Merge label across A–H
    ws.merge_cells(f"A{totals_row}:H{totals_row}")

    # Total Cost Basis sum
    tcb_total = ws.cell(
        row=totals_row, column=9,
        value=f"=SUM(I{data_start_row}:I{last_data_row})"
    )
    tcb_total.font         = Font(name="Calibri", bold=True, size=10)
    tcb_total.fill         = navy_fill
    tcb_total.alignment    = right_align
    tcb_total.border       = border
    tcb_total.number_format = '$#,##0.00'
    tcb_total.font         = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    # Total Market Value sum
    mv_total = ws.cell(
        row=totals_row, column=8,
        value=f"=IF(COUNTA(G{data_start_row}:G{last_data_row})=0,\"\","
              f"SUM(H{data_start_row}:H{last_data_row}))"
    )
    mv_total.font         = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    mv_total.fill         = navy_fill
    mv_total.alignment    = right_align
    mv_total.border       = border
    mv_total.number_format = '$#,##0.00'

    # Unmerge A–H and redo to leave col 8 and 9 separate
    ws.unmerge_cells(f"A{totals_row}:H{totals_row}")
    ws.merge_cells(f"A{totals_row}:G{totals_row}")

    for col in range(1, 13):
        c = ws.cell(row=totals_row, column=col)
        c.fill   = navy_fill
        c.border = border
        if c.font.color.rgb != "FFFFFF":
            c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    # ── Freeze panes & filters ───────────────────────────────
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:L{last_data_row}"

    # ── Entity summary sheet ─────────────────────────────────
    ws2 = wb.create_sheet("Entity Summary")
    ws2["A1"] = "Entity Summary"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill = navy_fill
    ws2["A1"].alignment = center_align
    ws2.merge_cells("A1:E1")
    ws2.row_dimensions[1].height = 24

    ent_hdrs = ["Entity", "# Positions", "Total Shares", "Total Cost Basis", "% of Portfolio"]
    for ci, h in enumerate(ent_hdrs, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = subhdr_font
        c.fill = PatternFill("solid", fgColor="EBF0F5")
        c.alignment = center_align
        c.border = border
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 16

    entity_groups = positions.groupby("Entity").agg(
        positions_count=("Ticker", "count"),
        total_shares=("Total Shares", "sum"),
        total_cost=("Total Cost Basis", "sum"),
    ).reset_index()
    grand_total = entity_groups["total_cost"].sum()

    for ri, (_, eg) in enumerate(entity_groups.iterrows(), start=3):
        ws2.cell(row=ri, column=1, value=eg["Entity"]).border = border
        ws2.cell(row=ri, column=2, value=eg["positions_count"]).border = border
        c3 = ws2.cell(row=ri, column=3, value=eg["total_shares"])
        c3.number_format = "#,##0.0000"; c3.border = border
        c4 = ws2.cell(row=ri, column=4, value=eg["total_cost"])
        c4.number_format = "$#,##0.00";  c4.border = border
        c5 = ws2.cell(row=ri, column=5, value=eg["total_cost"] / grand_total if grand_total else 0)
        c5.number_format = "0.00%"; c5.border = border
        for ci in range(1, 6):
            ws2.cell(row=ri, column=ci).font = data_font

    wb.save(str(EXCEL_PATH))
    print(f"  Excel saved → {EXCEL_PATH}")


# ══════════════════════════════════════════════════════════
# 6. MAIN
# ══════════════════════════════════════════════════════════

def main():
    print("\n" + "=" * 60)
    print("  WCCP Portfolio Analyzer")
    print("=" * 60)

    print("\n[1/5] Loading trade blotter …")
    blotter = load_blotter(BLOTTER_FILE)
    print(f"      {len(blotter)} trades loaded  "
          f"({blotter['Transaction Type'].value_counts().to_dict()})")

    print("\n[2/5] Running FIFO lot matching …")
    open_lots, realized_pl = fifo_match(blotter)
    print(f"      {len(open_lots)} open lots  |  {len(realized_pl)} matched sells")

    print("\n[3/5] Building position summary …")
    positions = build_positions(open_lots)
    print(f"      {len(positions)} open positions")
    print(f"      Total cost basis: ${positions['Total Cost Basis'].sum():,.2f}")

    if positions.empty:
        print("\n[WARN] No open positions found. Check your blotter data.")
        return

    print("\n[4/5] Generating PDF report …")
    generate_pdf(positions, open_lots, realized_pl, blotter)

    print("\n[5/5] Generating Excel pricing file …")
    generate_excel(positions)

    print("\n" + "=" * 60)
    print("  Done!")
    print(f"  PDF    → {PDF_PATH}")
    print(f"  Excel  → {EXCEL_PATH}")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
